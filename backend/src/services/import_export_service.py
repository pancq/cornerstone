"""数据导入导出服务"""
import io
import csv
from typing import List, Dict, Any, Callable, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from ..models import Device, IPAddress, Circuit, Site, Prefix
from ..utils.logger import audit_log


class ImportExportService:
    """导入导出服务基类"""
    
    @staticmethod
    async def export_to_excel(
        data: List[Dict[str, Any]],
        headers: Dict[str, str],
        filename: str = "export.xlsx"
    ) -> bytes:
        """
        导出数据到Excel
        
        Args:
            data: 数据列表
            headers: 表头映射 {字段名: 显示名称}
            filename: 文件名
        
        Returns:
            Excel文件的字节流
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "数据导出"
        
        # 写入表头
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, (field, display_name) in enumerate(headers.items(), start=1):
            cell = ws.cell(row=1, column=col_idx, value=display_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入数据
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, field in enumerate(headers.keys(), start=1):
                value = row_data.get(field, "")
                # 处理特殊类型
                if isinstance(value, bool):
                    value = "是" if value else "否"
                elif value is None:
                    value = ""
                ws.cell(row=row_idx, column=col_idx, value=str(value))
        
        # 自动调整列宽
        for col_idx, field in enumerate(headers.keys(), start=1):
            max_length = len(headers[field])
            for row_idx in range(2, len(data) + 2):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)
        
        # 保存到字节流
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    @staticmethod
    async def export_to_csv(
        data: List[Dict[str, Any]],
        headers: Dict[str, str]
    ) -> str:
        """
        导出数据到CSV
        
        Args:
            data: 数据列表
            headers: 表头映射 {字段名: 显示名称}
        
        Returns:
            CSV字符串
        """
        output = io.StringIO()
        writer = csv.writer(output)
        
        # 写入表头
        writer.writerow(list(headers.values()))
        
        # 写入数据
        for row_data in data:
            row = []
            for field in headers.keys():
                value = row_data.get(field, "")
                if isinstance(value, bool):
                    value = "是" if value else "否"
                elif value is None:
                    value = ""
                row.append(str(value))
            writer.writerow(row)
        
        return output.getvalue()
    
    @staticmethod
    async def import_from_excel(
        file_content: bytes,
        field_mapping: Dict[str, str],
        required_fields: List[str] = []
    ) -> List[Dict[str, Any]]:
        """
        从Excel导入数据
        
        Args:
            file_content: Excel文件字节流
            field_mapping: 字段映射 {Excel列名: 模型字段名}
            required_fields: 必填字段列表
        
        Returns:
            解析后的数据列表
        """
        wb = load_workbook(io.BytesIO(file_content))
        ws = wb.active
        
        # 读取表头
        headers = {}
        for col_idx in range(1, ws.max_column + 1):
            header_value = ws.cell(row=1, column=col_idx).value
            if header_value:
                headers[header_value.strip()] = col_idx
        
        # 验证必填字段
        missing_fields = []
        for field in required_fields:
            if field not in headers:
                missing_fields.append(field)
        
        if missing_fields:
            raise ValueError(f"缺少必填字段: {', '.join(missing_fields)}")
        
        # 读取数据
        data = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for excel_col, model_field in field_mapping.items():
                if excel_col in headers:
                    value = ws.cell(row=row_idx, column=headers[excel_col]).value
                    row_data[model_field] = value
            
            # 跳过空行
            if any(row_data.values()):
                data.append(row_data)
        
        return data
    
    @staticmethod
    async def import_from_csv(
        file_content: str,
        field_mapping: Dict[str, str],
        required_fields: List[str] = []
    ) -> List[Dict[str, Any]]:
        """
        从CSV导入数据
        
        Args:
            file_content: CSV字符串
            field_mapping: 字段映射 {CSV列名: 模型字段名}
            required_fields: 必填字段列表
        
        Returns:
            解析后的数据列表
        """
        reader = csv.DictReader(io.StringIO(file_content))
        
        # 验证必填字段
        if reader.fieldnames:
            missing_fields = []
            for field in required_fields:
                if field not in reader.fieldnames:
                    missing_fields.append(field)
            
            if missing_fields:
                raise ValueError(f"缺少必填字段: {', '.join(missing_fields)}")
        
        # 读取数据
        data = []
        for row in reader:
            row_data = {}
            for csv_col, model_field in field_mapping.items():
                value = row.get(csv_col, "")
                row_data[model_field] = value if value else None
            
            # 跳过空行
            if any(row_data.values()):
                data.append(row_data)
        
        return data


class PrefixImportExport(ImportExportService):
    """子网导入导出"""

    FIELD_MAPPING = {
        "子网": "network",
        "站点": "site_name",
        "VLAN": "vlan",
        "用途": "usage"
    }

    HEADERS = {
        "network": "子网",
        "site_name": "站点",
        "vlan": "VLAN",
        "usage": "用途",
        "created_at": "创建时间"
    }

    @staticmethod
    async def export_prefixes(db: AsyncSession, format: str = "excel") -> bytes:
        """导出子网列表"""
        result = await db.execute(select(Prefix))
        prefixes = result.scalars().all()

        data = []
        for prefix in prefixes:
            site_name = ""
            if prefix.site_id:
                site_result = await db.execute(select(Site).where(Site.id == prefix.site_id))
                site = site_result.scalar_one_or_none()
                site_name = site.name if site else ""

            data.append({
                "network": prefix.network,
                "site_name": site_name,
                "vlan": prefix.vlan,
                "usage": prefix.usage,
                "created_at": prefix.created_at.strftime("%Y-%m-%d %H:%M:%S") if prefix.created_at else ""
            })

        if format == "csv":
            csv_content = await ImportExportService.export_to_csv(data, PrefixImportExport.HEADERS)
            return csv_content.encode('utf-8-sig')
        else:
            return await ImportExportService.export_to_excel(data, PrefixImportExport.HEADERS)

    @staticmethod
    async def import_prefixes(
        db: AsyncSession,
        file_content: bytes,
        file_type: str,
        user: str
    ) -> Dict[str, Any]:
        """导入子网（覆盖重复）"""
        if file_type == "excel":
            data = await ImportExportService.import_from_excel(
                file_content,
                PrefixImportExport.FIELD_MAPPING,
                required_fields=["子网"]
            )
        else:
            data = await ImportExportService.import_from_csv(
                file_content.decode('utf-8-sig'),
                PrefixImportExport.FIELD_MAPPING,
                required_fields=["子网"]
            )

        success_count = 0
        failed_count = 0
        errors = []

        for idx, row_data in enumerate(data, start=2):
            try:
                network = row_data.get("network")
                if not network:
                    failed_count += 1
                    errors.append(f"第{idx}行: 子网不能为空")
                    continue

                site_id = None
                if row_data.get("site_name"):
                    site_result = await db.execute(
                        select(Site).where(Site.name == row_data["site_name"])
                    )
                    site = site_result.scalar_one_or_none()
                    site_id = site.id if site else None

                existing_result = await db.execute(
                    select(Prefix).where(Prefix.network == network)
                )
                existing_prefix = existing_result.scalar_one_or_none()

                if existing_prefix:
                    existing_prefix.site_id = site_id
                    existing_prefix.vlan = row_data.get("vlan")
                    existing_prefix.usage = row_data.get("usage")
                else:
                    prefix = Prefix(
                        network=network,
                        site_id=site_id,
                        vlan=row_data.get("vlan"),
                        usage=row_data.get("usage")
                    )
                    db.add(prefix)

                success_count += 1
            except Exception as e:
                failed_count += 1
                errors.append(f"第{idx}行: {str(e)}")

        await db.commit()

        await audit_log(
            db,
            user,
            "prefix",
            0,
            "批量导入子网",
            {"success_count": success_count, "failed_count": failed_count}
        )

        return {
            "success_count": success_count,
            "failed_count": failed_count,
            "errors": errors[:10]
        }


class DeviceImportExport(ImportExportService):
    """设备导入导出"""

    FIELD_MAPPING = {
        "设备名称": "name",
        "设备类型": "type",
        "IP地址": "ip_address",
        "厂商": "brand",
        "型号": "model",
        "站点": "site_name",
        "位置": "location",
        "状态": "status",
        "描述": "note"
    }
    
    HEADERS = {
        "name": "设备名称",
        "type": "设备类型",
        "ip_address": "IP地址",
        "brand": "厂商",
        "model": "型号",
        "site_name": "站点",
        "location": "位置",
        "status": "状态",
        "note": "描述",
        "created_at": "创建时间"
    }
    
    @staticmethod
    async def export_devices(db: AsyncSession, format: str = "excel") -> bytes:
        """导出设备列表"""
        result = await db.execute(select(Device))
        devices = result.scalars().all()
        
        data = []
        for device in devices:
            # 获取站点名称
            site_name = ""
            if device.site_id:
                site_result = await db.execute(select(Site).where(Site.id == device.site_id))
                site = site_result.scalar_one_or_none()
                site_name = site.name if site else ""
            
            # 获取管理IP
            ip_address = ""
            if device.mgmt_ip_id:
                ip_result = await db.execute(select(IPAddress).where(IPAddress.id == device.mgmt_ip_id))
                ip_addr = ip_result.scalar_one_or_none()
                ip_address = ip_addr.address if ip_addr else ""
            
            data.append({
                "name": device.name,
                "type": device.type,
                "ip_address": ip_address,
                "brand": device.brand,
                "model": device.model,
                "site_name": site_name,
                "location": device.location,
                "status": device.status,
                "note": device.note,
                "created_at": device.created_at.strftime("%Y-%m-%d %H:%M:%S") if device.created_at else ""
            })
        
        if format == "csv":
            csv_content = await ImportExportService.export_to_csv(data, DeviceImportExport.HEADERS)
            return csv_content.encode('utf-8-sig')  # 添加BOM以支持Excel打开
        else:
            return await ImportExportService.export_to_excel(data, DeviceImportExport.HEADERS)
    
    @staticmethod
    async def import_devices(
        db: AsyncSession,
        file_content: bytes,
        file_type: str,
        user: str
    ) -> Dict[str, Any]:
        """导入设备"""
        # 解析文件
        if file_type == "excel":
            data = await ImportExportService.import_from_excel(
                file_content,
                DeviceImportExport.FIELD_MAPPING,
                required_fields=["设备名称"]
            )
        else:
            data = await ImportExportService.import_from_csv(
                file_content.decode('utf-8-sig'),
                DeviceImportExport.FIELD_MAPPING,
                required_fields=["设备名称"]
            )
        
        # 导入数据
        success_count = 0
        failed_count = 0
        errors = []
        
        for idx, row_data in enumerate(data, start=2):  # Excel行号从2开始
            try:
                # 查找站点
                site_id = None
                if row_data.get("site_name"):
                    site_result = await db.execute(
                        select(Site).where(Site.name == row_data["site_name"])
                    )
                    site = site_result.scalar_one_or_none()
                    site_id = site.id if site else None
                
                # 创建设备
                device = Device(
                    name=row_data.get("name"),
                    type=row_data.get("type", "unknown"),
                    brand=row_data.get("brand"),
                    model=row_data.get("model"),
                    site_id=site_id,
                    location=row_data.get("location"),
                    status=row_data.get("status", "active"),
                    note=row_data.get("note")
                )
                db.add(device)
                success_count += 1
            except Exception as e:
                failed_count += 1
                errors.append(f"第{idx}行: {str(e)}")
        
        await db.commit()
        
        # 记录审计日志
        await audit_log(
            db,
            user,
            "device",
            0,
            "批量导入设备",
            {"success_count": success_count, "failed_count": failed_count}
        )
        
        return {
            "success_count": success_count,
            "failed_count": failed_count,
            "errors": errors[:10]  # 只返回前10条错误
        }


class IPAddressImportExport(ImportExportService):
    """IP地址导入导出"""

    FIELD_MAPPING = {
        "IP地址": "address",
        "子网": "prefix_name",
        "状态": "status",
        "用途": "usage",
        "负责人": "owner",
        "过期时间": "expire_at"
    }

    HEADERS = {
        "address": "IP地址",
        "prefix_name": "子网",
        "status": "状态",
        "usage": "用途",
        "owner": "负责人",
        "expire_at": "过期时间",
        "created_at": "创建时间"
    }

    @staticmethod
    async def export_ip_addresses(db: AsyncSession, prefix_id: Optional[int] = None, format: str = "excel") -> bytes:
        """导出IP地址列表"""
        query = select(IPAddress)
        if prefix_id:
            query = query.where(IPAddress.prefix_id == prefix_id)

        result = await db.execute(query)
        ip_addresses = result.scalars().all()

        data = []
        for ip in ip_addresses:
            # 获取子网名称
            prefix_name = ""
            if ip.prefix_id:
                from ..models import Prefix
                prefix_result = await db.execute(select(Prefix).where(Prefix.id == ip.prefix_id))
                prefix = prefix_result.scalar_one_or_none()
                prefix_name = prefix.network if prefix else ""

            data.append({
                "address": ip.address,
                "prefix_name": prefix_name,
                "status": ip.status,
                "usage": ip.usage,
                "owner": ip.owner,
                "expire_at": ip.expire_at.strftime("%Y-%m-%d") if ip.expire_at else "",
                "created_at": ip.created_at.strftime("%Y-%m-%d %H:%M:%S") if ip.created_at else ""
            })

        if format == "csv":
            csv_content = await ImportExportService.export_to_csv(data, IPAddressImportExport.HEADERS)
            return csv_content.encode('utf-8-sig')
        else:
            return await ImportExportService.export_to_excel(data, IPAddressImportExport.HEADERS)

    @staticmethod
    async def import_ip_addresses(
        db: AsyncSession,
        file_content: bytes,
        file_type: str,
        user: str,
        prefix_id: Optional[int] = None
    ) -> Dict[str, Any]:
        """导入IP地址"""
        # 解析文件
        if file_type == "excel":
            data = await ImportExportService.import_from_excel(
                file_content,
                IPAddressImportExport.FIELD_MAPPING,
                required_fields=["IP地址"]
            )
        else:
            data = await ImportExportService.import_from_csv(
                file_content.decode('utf-8-sig'),
                IPAddressImportExport.FIELD_MAPPING,
                required_fields=["IP地址"]
            )

        # 导入数据
        success_count = 0
        failed_count = 0
        errors = []

        for idx, row_data in enumerate(data, start=2):
            try:
                # 查找子网
                ip_prefix_id = prefix_id
                if not ip_prefix_id and row_data.get("prefix_name"):
                    from ..models import Prefix
                    prefix_result = await db.execute(
                        select(Prefix).where(Prefix.network == row_data["prefix_name"])
                    )
                    prefix = prefix_result.scalar_one_or_none()
                    ip_prefix_id = prefix.id if prefix else None

                # 解析过期时间
                expire_at = None
                if row_data.get("expire_at"):
                    from datetime import datetime
                    try:
                        expire_at = datetime.strptime(row_data["expire_at"], "%Y-%m-%d")
                    except:
                        pass

                # 创建IP地址
                ip_address = IPAddress(
                    address=row_data.get("address"),
                    prefix_id=ip_prefix_id,
                    status=row_data.get("status", "未分配"),
                    usage=row_data.get("usage"),
                    owner=row_data.get("owner"),
                    expire_at=expire_at
                )
                db.add(ip_address)
                success_count += 1
            except Exception as e:
                failed_count += 1
                errors.append(f"第{idx}行: {str(e)}")

        await db.commit()

        # 记录审计日志
        await audit_log(
            db,
            user,
            "ip_address",
            0,
            "批量导入IP地址",
            {"success_count": success_count, "failed_count": failed_count}
        )

        return {
            "success_count": success_count,
            "failed_count": failed_count,
            "errors": errors[:10]
        }


class CircuitImportExport(ImportExportService):
    """电路导入导出"""

    FIELD_MAPPING = {
        "电路名称": "name",
        "运营商": "provider",
        "类型": "type",
        "站点": "site_name",
        "带宽": "bandwidth",
        "月费": "monthly_cost",
        "合同开始": "contract_start",
        "合同结束": "contract_end",
        "电路编号": "circuit_no",
        "客服电话": "support_phone",
        "公网IP": "public_ip",
        "状态": "status",
        "备注": "note"
    }

    HEADERS = {
        "name": "电路名称",
        "provider": "运营商",
        "type": "类型",
        "site_name": "站点",
        "bandwidth": "带宽",
        "monthly_cost": "月费",
        "contract_start": "合同开始",
        "contract_end": "合同结束",
        "circuit_no": "电路编号",
        "support_phone": "客服电话",
        "public_ip": "公网IP",
        "status": "状态",
        "note": "备注",
        "updated_at": "更新时间"
    }

    @staticmethod
    async def export_circuits(db: AsyncSession, format: str = "excel") -> bytes:
        """导出电路列表"""
        result = await db.execute(select(Circuit))
        circuits = result.scalars().all()

        data = []
        for circuit in circuits:
            # 获取站点名称
            site_name = ""
            if circuit.site_id:
                site_result = await db.execute(select(Site).where(Site.id == circuit.site_id))
                site = site_result.scalar_one_or_none()
                site_name = site.name if site else ""

            data.append({
                "name": circuit.name,
                "provider": circuit.provider,
                "type": circuit.type,
                "site_name": site_name,
                "bandwidth": circuit.bandwidth,
                "monthly_cost": circuit.monthly_cost,
                "contract_start": circuit.contract_start.strftime("%Y-%m-%d") if circuit.contract_start else "",
                "contract_end": circuit.contract_end.strftime("%Y-%m-%d") if circuit.contract_end else "",
                "circuit_no": circuit.circuit_no,
                "support_phone": circuit.support_phone,
                "public_ip": circuit.public_ip,
                "status": circuit.status,
                "note": circuit.note,
                "updated_at": circuit.updated_at.strftime("%Y-%m-%d %H:%M:%S") if circuit.updated_at else ""
            })

        if format == "csv":
            csv_content = await ImportExportService.export_to_csv(data, CircuitImportExport.HEADERS)
            return csv_content.encode('utf-8-sig')
        else:
            return await ImportExportService.export_to_excel(data, CircuitImportExport.HEADERS)

    @staticmethod
    async def import_circuits(
        db: AsyncSession,
        file_content: bytes,
        file_type: str,
        user: str
    ) -> Dict[str, Any]:
        """导入电路"""
        # 解析文件
        if file_type == "excel":
            data = await ImportExportService.import_from_excel(
                file_content,
                CircuitImportExport.FIELD_MAPPING,
                required_fields=["电路名称"]
            )
        else:
            data = await ImportExportService.import_from_csv(
                file_content.decode('utf-8-sig'),
                CircuitImportExport.FIELD_MAPPING,
                required_fields=["电路名称"]
            )

        # 导入数据
        success_count = 0
        failed_count = 0
        errors = []

        for idx, row_data in enumerate(data, start=2):
            try:
                # 查找站点
                site_id = None
                if row_data.get("site_name"):
                    site_result = await db.execute(
                        select(Site).where(Site.name == row_data["site_name"])
                    )
                    site = site_result.scalar_one_or_none()
                    site_id = site.id if site else None

                # 解析日期
                from datetime import datetime
                contract_start = None
                contract_end = None
                if row_data.get("contract_start"):
                    try:
                        contract_start = datetime.strptime(row_data["contract_start"], "%Y-%m-%d")
                    except:
                        pass
                if row_data.get("contract_end"):
                    try:
                        contract_end = datetime.strptime(row_data["contract_end"], "%Y-%m-%d")
                    except:
                        pass

                # 创建电路
                circuit = Circuit(
                    name=row_data.get("name"),
                    provider=row_data.get("provider"),
                    type=row_data.get("type", "互联网专线"),
                    site_id=site_id,
                    bandwidth=row_data.get("bandwidth"),
                    monthly_cost=row_data.get("monthly_cost"),
                    contract_start=contract_start,
                    contract_end=contract_end,
                    circuit_no=row_data.get("circuit_no"),
                    support_phone=row_data.get("support_phone"),
                    public_ip=row_data.get("public_ip"),
                    status=row_data.get("status", "正常"),
                    note=row_data.get("note"),
                    updated_by=user
                )
                db.add(circuit)
                success_count += 1
            except Exception as e:
                failed_count += 1
                errors.append(f"第{idx}行: {str(e)}")

        await db.commit()

        # 记录审计日志
        await audit_log(
            db,
            user,
            "circuit",
            0,
            "批量导入电路",
            {"success_count": success_count, "failed_count": failed_count}
        )
        
        return {
            "success_count": success_count,
            "failed_count": failed_count,
            "errors": errors[:10]
        }

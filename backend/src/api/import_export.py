"""数据导入导出API"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
import io

from ..database import get_db
from ..models import User
from ..api.dependencies import get_current_active_user
from ..services.import_export_service import (
    DeviceImportExport,
    IPAddressImportExport,
    CircuitImportExport,
    PrefixImportExport
)

router = APIRouter()


# ==================== 设备导入导出 ====================

@router.get("/devices/export")
async def export_devices(
    format: str = Query("excel", description="导出格式: excel 或 csv"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """导出设备列表"""
    try:
        file_content = await DeviceImportExport.export_devices(db, format)
        
        if format == "csv":
            filename = "devices.csv"
            media_type = "text/csv"
        else:
            filename = "devices.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
        return StreamingResponse(
            io.BytesIO(file_content),
            media_type=media_type,
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


@router.post("/devices/import")
async def import_devices(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """导入设备"""
    # 检查文件类型
    filename = file.filename.lower()
    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        file_type = "excel"
    elif filename.endswith('.csv'):
        file_type = "csv"
    else:
        raise HTTPException(status_code=400, detail="不支持的文件格式，请上传Excel或CSV文件")
    
    # 读取文件内容
    file_content = await file.read()
    
    # 执行导入
    try:
        result = await DeviceImportExport.import_devices(
            db=db,
            file_content=file_content,
            file_type=file_type,
            user=current_user.username
        )
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")


@router.get("/devices/template")
async def download_device_template():
    """下载设备导入模板"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "设备导入模板"
    
    # 表头
    headers = ["设备名称", "设备类型", "IP地址", "厂商", "型号", "站点", "位置", "状态", "描述"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 示例数据
    example_data = [
        ["SW-DEMO-CORE-01", "switch", "192.0.2.1", "Cisco", "Catalyst 9300", "Demo Site A", "Demo Rack A-U12", "active", "Demo core switch"],
        ["RT-DEMO-WAN-01", "router", "192.0.2.2", "H3C", "MSR 5660", "Demo Site B", "Demo Rack C-U08", "active", "Demo WAN router"],
        ["FW-DEMO-EDGE-01", "firewall", "192.0.2.3", "Huawei", "USG6655E", "Demo Site A", "Demo Rack B-U06", "active", "Demo edge firewall"]
    ]
    
    for row_idx, row_data in enumerate(example_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 调整列宽
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 15
    
    # 保存到字节流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="device_import_template.xlsx"'
        }
    )


# ==================== 子网导入导出 ====================

@router.get("/prefixes/export")
async def export_prefixes(
    format: str = Query("excel", description="导出格式: excel 或 csv"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """导出子网列表"""
    try:
        file_content = await PrefixImportExport.export_prefixes(db, format)
        filename = f"subnets.{format}"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if format == "excel" else "text/csv"
        return StreamingResponse(io.BytesIO(file_content), media_type=media_type, headers={"Content-Disposition": f"attachment; filename={filename}"})
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


@router.get("/prefixes/template")
async def get_prefix_template():
    """下载子网导入模板"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "子网导入模板"

    headers = ["站点", "子网", "VLAN", "用途"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 设置列宽
    column_widths = [15, 15, 8, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    # 添加示例数据
    sample_data = [
        ["总部", "192.168.1.0/24", "10", "办公网"],
        ["分部A", "192.168.2.0/24", "20", "服务器"],
        ["分部B", "192.168.3.0/24", "30", "物联网"],
    ]

    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="prefix_import_template.xlsx"'
        }
    )


@router.post("/prefixes/import")
async def import_prefixes(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """导入子网"""
    filename = file.filename.lower()
    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        file_type = "excel"
    elif filename.endswith('.csv'):
        file_type = "csv"
    else:
        raise HTTPException(status_code=400, detail="不支持的文件格式，请上传Excel或CSV文件")

    file_content = await file.read()

    try:
        result = await PrefixImportExport.import_prefixes(db=db, file_content=file_content, file_type=file_type, user=current_user.username)
        return {"success_count": result["success_count"], "failed_count": result["failed_count"], "errors": result["errors"]}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")


# ==================== IP地址导入导出 ====================

@router.get("/ip-addresses/export")
async def export_ip_addresses(
    prefix_id: int = Query(None, description="子网ID，不传则导出所有"),
    format: str = Query("excel", description="导出格式: excel 或 csv"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """导出IP地址列表"""
    try:
        file_content = await IPAddressImportExport.export_ip_addresses(db, prefix_id, format)
        
        if format == "csv":
            filename = "ip_addresses.csv"
            media_type = "text/csv"
        else:
            filename = "ip_addresses.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
        return StreamingResponse(
            io.BytesIO(file_content),
            media_type=media_type,
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


@router.post("/ip-addresses/import")
async def import_ip_addresses(
    file: UploadFile = File(...),
    prefix_id: int = Query(None, description="子网ID"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """导入IP地址"""
    # 检查文件类型
    filename = file.filename.lower()
    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        file_type = "excel"
    elif filename.endswith('.csv'):
        file_type = "csv"
    else:
        raise HTTPException(status_code=400, detail="不支持的文件格式，请上传Excel或CSV文件")
    
    # 读取文件内容
    file_content = await file.read()
    
    # 执行导入
    try:
        result = await IPAddressImportExport.import_ip_addresses(
            db=db,
            file_content=file_content,
            file_type=file_type,
            user=current_user.username,
            prefix_id=prefix_id
        )
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")


@router.get("/ip-addresses/template")
async def download_ip_address_template():
    """下载IP地址导入模板"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "IP地址导入模板"
    
    # 表头
    headers = ["IP地址", "子网", "状态", "用途", "负责人", "描述", "过期时间"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 示例数据
    example_data = [
        ["192.0.2.10", "192.0.2.0/24", "已分配", "服务器", "Demo Owner A", "Demo web server", "2025-12-31"],
        ["192.0.2.11", "192.0.2.0/24", "已分配", "数据库", "Demo Owner B", "Demo database", "2025-12-31"],
        ["192.0.2.100", "192.0.2.0/24", "预留", "测试", "Demo Owner C", "Demo test environment", ""]
    ]
    
    for row_idx, row_data in enumerate(example_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 调整列宽
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 15
    
    # 保存到字节流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="ip_address_import_template.xlsx"'
        }
    )


# ==================== 电路导入导出 ====================

@router.get("/circuits/export")
async def export_circuits(
    format: str = Query("excel", description="导出格式: excel 或 csv"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """导出电路列表"""
    try:
        file_content = await CircuitImportExport.export_circuits(db, format)
        
        if format == "csv":
            filename = "circuits.csv"
            media_type = "text/csv"
        else:
            filename = "circuits.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
        return StreamingResponse(
            io.BytesIO(file_content),
            media_type=media_type,
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


@router.post("/circuits/import")
async def import_circuits(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """导入电路"""
    # 检查文件类型
    filename = file.filename.lower()
    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        file_type = "excel"
    elif filename.endswith('.csv'):
        file_type = "csv"
    else:
        raise HTTPException(status_code=400, detail="不支持的文件格式，请上传Excel或CSV文件")
    
    # 读取文件内容
    file_content = await file.read()
    
    # 执行导入
    try:
        result = await CircuitImportExport.import_circuits(
            db=db,
            file_content=file_content,
            file_type=file_type,
            user=current_user.username
        )
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")


@router.get("/circuits/template")
async def download_circuit_template():
    """下载电路导入模板"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "电路导入模板"
    
    # 表头
    headers = [
        "电路编号", "电路名称", "类型", "带宽", "起点站点", "起点设备", "起点端口",
        "终点站点", "终点设备", "终点端口", "运营商", "状态", "到期时间", "描述"
    ]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 示例数据
    example_data = [
        ["CIR-001", "北京-上海专线", "mpls", "100M", "北京机房", "核心交换机-01", "Gig0/1", 
         "上海机房", "核心交换机-02", "Gig0/1", "中国电信", "active", "2025-12-31", "跨省专线"],
        ["CIR-002", "总部-分部互联", "vpn", "50M", "总部机房", "边界路由器-01", "Gig0/2",
         "分部机房", "边界路由器-02", "Gig0/2", "中国移动", "active", "2025-06-30", "VPN互联"]
    ]
    
    for row_idx, row_data in enumerate(example_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 调整列宽
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 15
    
    # 保存到字节流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="circuit_import_template.xlsx"'
        }
    )
